"""Stage 5 - EXPORT. Writes assigned, QA-passed images into the Amazon
SAREE flat file Template tab. The xlsm is generated/updated per batch from
the SQLite registry, so the workbook itself never has to be the growing master.

Template layout (new Amazon.in template, verified from your SAREE.xlsm):
  row 4 (1-idx)  = machine field names, data starts row 6
  col A=SKU  B=Product Type  G=Item Name
  col S=Main Image URL, T..AA = Other Image URL 1-8, AB = Swatch
"""
from openpyxl import load_workbook

COL = {"SKU": 1, "PTYPE": 2, "ACTION": 3, "ITEM_NAME": 7,
       "MAIN": 19, "OTHER1": 20, "SWATCH": 28}   # OTHER1..8 = 20..27
DATA_START = 6

def slot_col(slot):
    if slot == "MAIN":   return COL["MAIN"]
    if slot == "SWATCH": return COL["SWATCH"]
    if slot.startswith("OTHER"):
        n = int(slot[5:] or 1)
        return COL["OTHER1"] + n - 1
    raise ValueError(slot)

def export_batch(template_path, out_path, rows, url_base):
    """rows: list of {sku, item_name, item_type, images: {slot: lib_path}}.
    url_base: where your images will be hosted / or use Amazon image upload tool
    and leave URL columns pointing at final hosted names."""
    wb = load_workbook(template_path, keep_vba=True)   # keep_vba: macros survive
    ws = wb["Template"]

    saree_items = [i for i in rows if i["item_type"] == "saree"]
    skipped = [i for i in rows if i["item_type"] != "saree"]
    for i in skipped:
        print(f"  [SKIP] {i['sku']} is '{i['item_type']}' - this is a SAREE template;"
              f" export it with its own flat file (kept ready in registry)")

    def row_free(r):
        v = ws.cell(row=r, column=COL["SKU"]).value
        return v is None or str(v).strip() == ""

    r, written = DATA_START, 0
    for item in saree_items:
        while not row_free(r):          # skips Amazon example row + note row
            r += 1
        ws.cell(row=r, column=COL["SKU"], value=item["sku"])
        ws.cell(row=r, column=COL["PTYPE"], value="SAREE")
        ws.cell(row=r, column=COL["ITEM_NAME"], value=item.get("item_name", ""))
        for slot, libpath in item["images"].items():
            fname = libpath.replace("\\", "/").split("/")[-1]
            ws.cell(row=r, column=slot_col(slot), value=f"{url_base.rstrip('/')}/{fname}")
        r += 1; written += 1
    wb.save(out_path)
    return out_path, written, [i["sku"] for i in skipped]
