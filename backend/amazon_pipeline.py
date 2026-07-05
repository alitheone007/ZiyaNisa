"""
Amazon flat-file pipeline — admin-only API.

Ported from the saree-pipeline CLI into the ZiyaNisa backend:
  intake  → QA (PIL, free) + dedupe (perceptual hash, free) + register in Mongo
  classify→ vision AI (pluggable provider: claude / ollama), confidence-gated
  assign  → PATCH per image: item type, SKU, slot, QA override
  export  → write assigned+passed rows into any uploaded Amazon .xlsm template

Design rules carried over from the CLI:
  - Mongo (az_images / az_skus / az_templates) is the master; each xlsm is a
    small per-batch upload file, never the growing database.
  - Images are classified at most once (ai_checked flag = permanent cache).
  - Low-confidence AI answers become item_type "?" and land in manual review
    instead of being guessed — that is how published accuracy stays >90%.
  - Templates are introspected from Amazon's own machine-name row (row 5), so
    any future flat-file template (dress material, lehenga, ...) plugs in
    without code changes.
"""
import asyncio
import base64
import hashlib
import io
import json
import os
import re
import time
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import List, Optional

import httpx
from fastapi import APIRouter, File, Form, Header, HTTPException, UploadFile
from pydantic import BaseModel

from PIL import Image, ImageFilter, ImageStat

try:
    import imagehash
    _IMAGEHASH_OK = True
except ImportError:
    _IMAGEHASH_OK = False

try:
    from openpyxl import load_workbook
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False

AMAZON_DIR   = Path("/app/uploads/amazon")
LIBRARY_DIR  = AMAZON_DIR / "library"
THUMBS_DIR   = AMAZON_DIR / "thumbs"
TEMPLATE_DIR = AMAZON_DIR / "templates"
EXPORT_DIR   = AMAZON_DIR / "exports"
for _d in (LIBRARY_DIR, THUMBS_DIR, TEMPLATE_DIR, EXPORT_DIR):
    _d.mkdir(parents=True, exist_ok=True)

# ── QA thresholds (Amazon.in apparel image rules) ─────────────────────────────
MIN_LONG_SIDE_HARD = 500
MIN_LONG_SIDE_ZOOM = 1000
MAX_ASPECT         = 2.2
BLUR_THRESHOLD     = 45.0
HAMMING_DUP        = 6      # <=6 differing bits on 64-bit phash => same photo

SLOTS = ["MAIN"] + [f"OTHER{i}" for i in range(1, 9)] + ["SWATCH"]

ITEM_TYPES = ["saree", "dress_material", "lehenga", "kurti", "blouse", "dupatta", "other", "?"]

CLASSIFY_PROMPT = """You are classifying Indian garment product photos for an Amazon listing pipeline.
For EACH numbered image, return a JSON array (same order) of objects:
{"n": <number>, "item_type": "saree"|"dress_material"|"lehenga"|"kurti"|"blouse"|"dupatta"|"other",
 "confidence": "high"|"low",
 "color": "<dominant color>", "fabric_guess": "<silk/cotton/georgette/chiffon/banarasi/etc or unknown>",
 "listing_quality": "good"|"poor",
 "issues": "<short note if poor, else empty>"}
If unsure of item_type, use confidence "low" - do NOT guess. Reply ONLY with the JSON array."""


# ── Pure-Python image helpers (run in thread pool — PIL is sync) ──────────────

def _blur_score(img):
    g = img.convert("L")
    g.thumbnail((800, 800))
    edges = g.filter(ImageFilter.FIND_EDGES)
    return ImageStat.Stat(edges).var[0]


def _white_bg_ratio(img):
    im = img.convert("RGB").resize((200, 200))
    px = im.load()
    border, white = 0, 0
    for i in range(200):
        for j in (0, 1, 198, 199):
            for p in (px[i, j], px[j, i]):
                border += 1
                if min(p) > 235:
                    white += 1
    return white / border


def _qa_check_bytes(data: bytes, size_bytes: int) -> dict:
    reasons, status = [], "pass"
    try:
        img = Image.open(io.BytesIO(data))
        img.load()
    except Exception as e:
        return {"status": "fail", "reasons": [f"unreadable: {e}"], "main_ok": False, "size": ""}

    w, h = img.size
    long_side = max(w, h)
    if long_side < MIN_LONG_SIDE_HARD:
        status = "fail"; reasons.append(f"too small {w}x{h} (<{MIN_LONG_SIDE_HARD}px) — re-request from supplier")
    elif long_side < MIN_LONG_SIDE_ZOOM:
        status = "manual_review"; reasons.append(f"below 1000px ({w}x{h}) — no zoom, usable as OTHER image only")

    ar = max(w, h) / max(1, min(w, h))
    if ar > MAX_ASPECT:
        status = "fail"; reasons.append(f"bad aspect ratio {ar:.1f} (likely screenshot/status crop)")

    if size_bytes > 10 * 1024 * 1024:
        if status != "fail":
            status = "manual_review"
        reasons.append("file >10MB — recompress")

    if img.format not in ("JPEG", "PNG", "WEBP"):
        if status != "fail":
            status = "manual_review"
        reasons.append(f"format {img.format} — convert to JPEG")

    blur = _blur_score(img)
    if blur < BLUR_THRESHOLD and status != "fail":
        status = "manual_review"; reasons.append(f"possibly blurry (score {blur:.0f})")

    wbg = _white_bg_ratio(img)
    main_ok = wbg > 0.85 and status == "pass"
    if not main_ok and status == "pass":
        reasons.append(f"background not pure white ({wbg:.0%}) — OK for OTHER slots, not MAIN")

    return {"status": status, "reasons": reasons, "main_ok": main_ok,
            "size": f"{w}x{h}", "white_bg": round(wbg, 2), "blur": round(blur, 1),
            "pixels": w * h}


def _hashes_bytes(data: bytes):
    img = Image.open(io.BytesIO(data)).convert("RGB")
    ph = str(imagehash.phash(img))
    sha = hashlib.sha1(data).hexdigest()
    return ph, sha


def _hamming(a_hex: str, b_hex: str) -> int:
    try:
        return bin(int(a_hex, 16) ^ int(b_hex, 16)).count("1")
    except ValueError:
        return 64


def _normalize_bytes(data: bytes, out_path: Path, max_side=3000):
    """sRGB JPEG q90 upload-ready copy + 320px thumb for the admin grid."""
    img = Image.open(io.BytesIO(data)).convert("RGB")
    if max(img.size) > max_side:
        img.thumbnail((max_side, max_side), Image.LANCZOS)
    img.save(out_path, "JPEG", quality=90, optimize=True)
    thumb = img.copy()
    thumb.thumbnail((320, 320), Image.LANCZOS)
    thumb.save(THUMBS_DIR / out_path.name, "JPEG", quality=80)


def _thumb_b64_from_file(path: Path, side=512) -> str:
    img = Image.open(path).convert("RGB")
    img.thumbnail((side, side), Image.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, "JPEG", quality=80)
    return base64.b64encode(buf.getvalue()).decode()


def _today() -> str:
    return datetime.now(timezone.utc).date().isoformat()


# ── Template introspection ─────────────────────────────────────────────────────
# Amazon flat files carry machine field names in row 5. We locate the columns
# we write by machine-name pattern, so ANY product-type template works.

MACHINE_ROW = 5
DATA_START  = 6

def _introspect_template(path: Path) -> dict:
    wb = load_workbook(path, read_only=True)
    if "Template" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="No 'Template' sheet — is this an Amazon flat file?")
    ws = wb["Template"]
    machine = next(ws.iter_rows(min_row=MACHINE_ROW, max_row=MACHINE_ROW, values_only=True))
    display = next(ws.iter_rows(min_row=4, max_row=4, values_only=True))
    cols = {}
    for idx, name in enumerate(machine, 1):
        if not name:
            continue
        n = str(name)
        if n.startswith("contribution_sku"):
            cols["sku"] = idx
        elif n.startswith("product_type"):
            cols["product_type"] = idx
        elif "record_action" in n:
            cols["action"] = idx
        elif n.startswith("item_name"):
            cols["item_name"] = idx
        elif n.startswith("main_product_image_locator"):
            cols["MAIN"] = idx
        elif n.startswith("other_product_image_locator"):
            m = re.search(r"_(\d+)", n.split("[")[0])
            if m:
                cols[f"OTHER{m.group(1)}"] = idx
        elif n.startswith("swatch_product_image_locator"):
            cols["SWATCH"] = idx
    wb.close()
    if "sku" not in cols or "MAIN" not in cols:
        raise HTTPException(status_code=400,
                            detail="Could not find SKU / Main Image columns in the Template sheet")
    total_cols = sum(1 for v in machine if v)
    return {"columns": cols, "total_columns": total_cols,
            "display_headers": [str(d) for d in display if d][:20]}


# ── Vision providers (pluggable) ───────────────────────────────────────────────

async def _classify_claude(paths: List[Path]) -> List[dict]:
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        raise HTTPException(status_code=400, detail="ANTHROPIC_API_KEY not set on the server")
    out: List[dict] = []
    async with httpx.AsyncClient(timeout=90.0) as client:
        for i in range(0, len(paths), 8):
            chunk = paths[i:i + 8]
            content = []
            for n, p in enumerate(chunk, 1):
                b64 = await asyncio.to_thread(_thumb_b64_from_file, p)
                content.append({"type": "text", "text": f"Image {n}:"})
                content.append({"type": "image", "source": {
                    "type": "base64", "media_type": "image/jpeg", "data": b64}})
            content.append({"type": "text", "text": CLASSIFY_PROMPT})
            resp = await client.post(
                "https://api.anthropic.com/v1/messages",
                headers={"x-api-key": api_key, "anthropic-version": "2023-06-01",
                         "content-type": "application/json"},
                json={"model": "claude-haiku-4-5-20251001", "max_tokens": 1500,
                      "messages": [{"role": "user", "content": content}]},
            )
            if resp.status_code != 200:
                raise HTTPException(status_code=502,
                                    detail=f"Claude API error {resp.status_code}: {resp.text[:200]}")
            txt = resp.json()["content"][0]["text"].strip()
            txt = txt.removeprefix("```json").removeprefix("```").removesuffix("```").strip()
            try:
                out.extend(json.loads(txt))
            except json.JSONDecodeError:
                out.extend([{"n": n, "item_type": "?", "confidence": "low",
                             "issues": "parse_error"} for n in range(1, len(chunk) + 1)])
    return out


async def _classify_ollama(paths: List[Path]) -> List[dict]:
    host  = os.environ.get("OLLAMA_HOST", "http://localhost:11434")
    model = os.environ.get("VISION_MODEL", "moondream")
    out: List[dict] = []
    async with httpx.AsyncClient(timeout=120.0) as client:
        for p in paths:   # local models: one image per call
            b64 = await asyncio.to_thread(_thumb_b64_from_file, p)
            try:
                resp = await client.post(f"{host}/api/generate", json={
                    "model": model, "prompt": CLASSIFY_PROMPT + "\nThere is exactly 1 image (n=1).",
                    "images": [b64], "stream": False})
                raw = resp.json().get("response", "").strip()
                raw = raw.removeprefix("```json").removeprefix("```").removesuffix("```").strip()
                parsed = json.loads(raw)
                out.append(parsed[0] if isinstance(parsed, list) else parsed)
            except Exception as exc:
                out.append({"item_type": "?", "confidence": "low", "issues": f"ollama: {exc}"})
    return out


VISION_PROVIDERS = {"claude": _classify_claude, "ollama": _classify_ollama}


# ── Google Drive intake (service account, no user OAuth) ──────────────────────

_GOOGLE_TOKEN_CACHE: dict = {}

async def _gdrive_token() -> str:
    sa_file = os.environ.get("GDRIVE_SA_FILE", "/app/gdrive-sa.json")
    if not os.path.exists(sa_file):
        raise HTTPException(status_code=400,
                            detail="Drive service-account file not found on server (GDRIVE_SA_FILE)")
    if _GOOGLE_TOKEN_CACHE.get("exp", 0) > time.time() + 60:
        return _GOOGLE_TOKEN_CACHE["token"]
    with open(sa_file) as f:
        sa = json.load(f)
    from jose import jwt as jose_jwt
    now = int(time.time())
    assertion = jose_jwt.encode(
        {"iss": sa["client_email"], "scope": "https://www.googleapis.com/auth/drive.readonly",
         "aud": sa["token_uri"], "iat": now, "exp": now + 3600},
        sa["private_key"], algorithm="RS256")
    async with httpx.AsyncClient(timeout=30.0) as client:
        resp = await client.post(sa["token_uri"], data={
            "grant_type": "urn:ietf:params:oauth:grant-type:jwt-bearer",
            "assertion": assertion})
    if resp.status_code != 200:
        raise HTTPException(status_code=502, detail=f"Google token error: {resp.text[:200]}")
    tok = resp.json()
    _GOOGLE_TOKEN_CACHE.update({"token": tok["access_token"], "exp": time.time() + tok.get("expires_in", 3600)})
    return tok["access_token"]


def _parse_drive_folder_id(s: str) -> str:
    s = s.strip()
    m = re.search(r"/folders/([A-Za-z0-9_-]{10,})", s)
    if m:
        return m.group(1)
    m = re.search(r"[?&]id=([A-Za-z0-9_-]{10,})", s)
    if m:
        return m.group(1)
    return s   # assume raw ID


async def _gdrive_list_images(folder: str, token: str, max_files=500) -> List[dict]:
    """BFS through folder + subfolders, return image file dicts."""
    images, queue, seen = [], [folder], set()
    async with httpx.AsyncClient(timeout=60.0) as client:
        while queue and len(images) < max_files:
            fid = queue.pop(0)
            if fid in seen:
                continue
            seen.add(fid)
            page = None
            while True:
                params = {
                    "q": f"'{fid}' in parents and trashed=false",
                    "fields": "nextPageToken, files(id,name,mimeType,size)",
                    "pageSize": 200, "supportsAllDrives": "true",
                    "includeItemsFromAllDrives": "true",
                }
                if page:
                    params["pageToken"] = page
                r = await client.get("https://www.googleapis.com/drive/v3/files",
                                     params=params, headers={"Authorization": f"Bearer {token}"})
                if r.status_code != 200:
                    raise HTTPException(status_code=502, detail=f"Drive list error: {r.text[:200]}")
                data = r.json()
                for f in data.get("files", []):
                    if f["mimeType"] == "application/vnd.google-apps.folder":
                        queue.append(f["id"])
                    elif f["mimeType"].startswith("image/"):
                        images.append(f)
                page = data.get("nextPageToken")
                if not page or len(images) >= max_files:
                    break
    return images[:max_files]


# ── Router factory ─────────────────────────────────────────────────────────────

class ImagePatch(BaseModel):
    item_type: Optional[str] = None
    sku: Optional[str] = None
    slot: Optional[str] = None
    qa_status: Optional[str] = None
    item_name: Optional[str] = None   # used when creating a new SKU


class ExportRequest(BaseModel):
    template_id: str
    url_base: str
    batch_date: Optional[str] = None   # limit to one batch; None = all ready


class DriveSyncRequest(BaseModel):
    folder: str


class ClassifyRequest(BaseModel):
    provider: str = "claude"


def make_amazon_router(db, token_from_header, is_admin_claims) -> APIRouter:
    router = APIRouter(prefix="/admin/amazon", tags=["amazon"])

    def guard(authorization: Optional[str]):
        claims = token_from_header(authorization)
        if not is_admin_claims(claims):
            raise HTTPException(status_code=403, detail="Admin access required")
        return claims

    async def _register_image_bytes(data: bytes, src_name: str, batch: str) -> dict:
        """QA + dedupe + store one image. Returns a result row for the UI."""
        if not _IMAGEHASH_OK:
            raise HTTPException(status_code=500, detail="imagehash not installed on server")
        try:
            ph, sha = await asyncio.to_thread(_hashes_bytes, data)
        except Exception as e:
            return {"file": src_name, "result": "unreadable", "detail": str(e)}

        # exact-byte duplicate
        hit = await db.az_images.find_one({"sha1": sha}, {"_id": 0})
        dist = 0
        if not hit:
            # near-duplicate (perceptual)
            async for row in db.az_images.find({}, {"_id": 0, "phash": 1, "sku": 1, "uploaded": 1,
                                                    "lib_file": 1, "pixels": 1}):
                d = _hamming(ph, row["phash"])
                if d <= HAMMING_DUP:
                    hit, dist = row, d
                    break
        if hit:
            qa = await asyncio.to_thread(_qa_check_bytes, data, len(data))
            # same photo, meaningfully higher resolution, not yet exported → upgrade
            if not hit.get("uploaded") and qa["status"] != "fail" \
                    and qa.get("pixels", 0) > (hit.get("pixels") or 0) * 1.2:
                lib_file = hit.get("lib_file") or f"{hit['phash']}.jpg"
                await asyncio.to_thread(_normalize_bytes, data, LIBRARY_DIR / lib_file)
                await db.az_images.update_one({"phash": hit["phash"]}, {"$set": {
                    "qa_status": qa["status"], "qa_reasons": "; ".join(qa["reasons"]),
                    "main_ok": qa["main_ok"], "size": qa["size"], "pixels": qa.get("pixels", 0),
                    "src_name": src_name}})
                return {"file": src_name, "result": "upgraded",
                        "detail": f"same photo, higher quality — library copy replaced (SKU: {hit.get('sku') or 'unassigned'})"}
            return {"file": src_name, "result": "duplicate",
                    "detail": f"already registered (dist {dist}, SKU: {hit.get('sku') or 'unassigned'})"}

        qa = await asyncio.to_thread(_qa_check_bytes, data, len(data))
        lib_file = ""
        if qa["status"] != "fail":
            lib_file = f"{ph}.jpg"
            await asyncio.to_thread(_normalize_bytes, data, LIBRARY_DIR / lib_file)
        await db.az_images.insert_one({
            "phash": ph, "sha1": sha, "src_name": src_name, "lib_file": lib_file,
            "batch_date": batch, "qa_status": qa["status"],
            "qa_reasons": "; ".join(qa["reasons"]), "main_ok": qa["main_ok"],
            "size": qa["size"], "pixels": qa.get("pixels", 0),
            "item_type": "", "ai_meta": None, "sku": "", "slot": "",
            "uploaded": False, "ai_checked": False,
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
        return {"file": src_name, "result": qa["status"],
                "detail": "; ".join(qa["reasons"]) or "passed all checks",
                "main_ok": qa["main_ok"]}

    # ── Intake ────────────────────────────────────────────────────────────────
    @router.post("/images")
    async def upload_images(files: List[UploadFile] = File(...),
                            authorization: Optional[str] = Header(None)):
        guard(authorization)
        batch = _today()
        results = []
        for f in files:
            data = await f.read()
            if len(data) > 25 * 1024 * 1024:
                results.append({"file": f.filename, "result": "unreadable", "detail": ">25MB"})
                continue
            results.append(await _register_image_bytes(data, f.filename or "upload", batch))
        summary = {}
        for r in results:
            summary[r["result"]] = summary.get(r["result"], 0) + 1
        return {"batch_date": batch, "results": results, "summary": summary}

    @router.post("/drive-sync")
    async def drive_sync(body: DriveSyncRequest,
                         authorization: Optional[str] = Header(None)):
        guard(authorization)
        folder = _parse_drive_folder_id(body.folder)
        token = await _gdrive_token()
        files = await _gdrive_list_images(folder, token)
        if not files:
            return {"batch_date": _today(), "results": [], "summary": {},
                    "note": "No images found — is the folder shared with the service account email?"}
        batch = _today()
        results = []
        async with httpx.AsyncClient(timeout=120.0) as client:
            for f in files:
                # skip files whose name we already ingested this batch is not enough —
                # dedupe by content happens in _register_image_bytes anyway
                r = await client.get(
                    f"https://www.googleapis.com/drive/v3/files/{f['id']}",
                    params={"alt": "media", "supportsAllDrives": "true"},
                    headers={"Authorization": f"Bearer {token}"})
                if r.status_code != 200:
                    results.append({"file": f["name"], "result": "unreadable",
                                    "detail": f"download failed ({r.status_code})"})
                    continue
                results.append(await _register_image_bytes(r.content, f["name"], batch))
        summary = {}
        for r in results:
            summary[r["result"]] = summary.get(r["result"], 0) + 1
        return {"batch_date": batch, "results": results, "summary": summary}

    # ── List / review ─────────────────────────────────────────────────────────
    @router.get("/images")
    async def list_images(status: Optional[str] = None, batch: Optional[str] = None,
                          assigned: Optional[str] = None, sku: Optional[str] = None,
                          uploaded: Optional[str] = None, limit: int = 200,
                          authorization: Optional[str] = Header(None)):
        guard(authorization)
        q = {}
        if status:
            q["qa_status"] = status
        if batch:
            q["batch_date"] = batch
        if sku:
            q["sku"] = sku
        if assigned == "yes":
            q["sku"] = {"$ne": ""}
        elif assigned == "no":
            q["sku"] = ""
        if uploaded == "yes":
            q["uploaded"] = True
        elif uploaded == "no":
            q["uploaded"] = False
        rows = await db.az_images.find(q, {"_id": 0}).sort("created_at", -1).to_list(min(limit, 500))
        return rows

    @router.patch("/images/{phash}")
    async def patch_image(phash: str, body: ImagePatch,
                          authorization: Optional[str] = Header(None)):
        guard(authorization)
        row = await db.az_images.find_one({"phash": phash}, {"_id": 0})
        if not row:
            raise HTTPException(status_code=404, detail="Image not found")
        updates = {}
        if body.item_type is not None:
            if body.item_type not in ITEM_TYPES and body.item_type != "":
                raise HTTPException(status_code=400, detail=f"item_type must be one of {ITEM_TYPES}")
            updates["item_type"] = body.item_type
        if body.qa_status is not None:
            if body.qa_status not in ("pass", "fail", "manual_review"):
                raise HTTPException(status_code=400, detail="qa_status must be pass/fail/manual_review")
            updates["qa_status"] = body.qa_status
        if body.sku is not None:
            updates["sku"] = body.sku.strip()
            if body.sku.strip():
                slot = (body.slot or "").strip().upper()
                if slot and slot not in SLOTS:
                    raise HTTPException(status_code=400, detail=f"slot must be one of {SLOTS}")
                taken = await db.az_images.find(
                    {"sku": body.sku.strip(), "phash": {"$ne": phash}, "slot": {"$ne": ""}},
                    {"_id": 0, "slot": 1}).to_list(20)
                used = {t["slot"] for t in taken}
                if not slot:   # auto-pick first free slot
                    free = [s for s in SLOTS if s not in used]
                    if not free:
                        raise HTTPException(status_code=409, detail="All 10 slots taken on this SKU")
                    slot = free[0]
                elif slot in used:
                    raise HTTPException(status_code=409,
                                        detail=f"Slot {slot} already taken on SKU {body.sku} (used: {sorted(used)})")
                updates["slot"] = slot
                # upsert SKU record
                itype = updates.get("item_type") or row.get("item_type") or "?"
                existing_sku = await db.az_skus.find_one({"sku": body.sku.strip()})
                if not existing_sku:
                    await db.az_skus.insert_one({
                        "sku": body.sku.strip(), "item_name": (body.item_name or "").strip(),
                        "item_type": itype, "created": _today()})
                elif body.item_name:
                    await db.az_skus.update_one({"sku": body.sku.strip()},
                                                {"$set": {"item_name": body.item_name.strip()}})
            else:
                updates["slot"] = ""
        await db.az_images.update_one({"phash": phash}, {"$set": updates})
        return {**row, **updates}

    @router.delete("/images/{phash}")
    async def delete_image(phash: str, authorization: Optional[str] = Header(None)):
        guard(authorization)
        row = await db.az_images.find_one({"phash": phash}, {"_id": 0})
        if not row:
            raise HTTPException(status_code=404, detail="Image not found")
        for d in (LIBRARY_DIR, THUMBS_DIR):
            p = d / (row.get("lib_file") or f"{phash}.jpg")
            if p.exists():
                p.unlink()
        await db.az_images.delete_one({"phash": phash})
        return {"deleted": phash}

    # ── AI classify ───────────────────────────────────────────────────────────
    @router.post("/classify")
    async def classify(body: ClassifyRequest,
                       authorization: Optional[str] = Header(None)):
        guard(authorization)
        provider = VISION_PROVIDERS.get(body.provider)
        if not provider:
            raise HTTPException(status_code=400,
                                detail=f"Unknown provider — available: {list(VISION_PROVIDERS)}")
        rows = await db.az_images.find(
            {"ai_checked": False, "qa_status": {"$ne": "fail"}, "lib_file": {"$ne": ""}},
            {"_id": 0}).to_list(200)
        if not rows:
            return {"classified": 0, "results": []}
        paths = [LIBRARY_DIR / r["lib_file"] for r in rows]
        results = await provider(paths)
        out = []
        for r, res in zip(rows, results):
            it = res.get("item_type", "?")
            if res.get("confidence") == "low":
                it = "?"
            sets = {"item_type": it if not r.get("item_type") else r["item_type"],
                    "ai_meta": {k: v for k, v in res.items() if k != "n"},
                    "ai_checked": True}
            if res.get("listing_quality") == "poor" and r["qa_status"] == "pass":
                sets["qa_status"] = "manual_review"
                sets["qa_reasons"] = (r.get("qa_reasons") or "") + f"; AI: {res.get('issues', 'poor listing quality')}"
            await db.az_images.update_one({"phash": r["phash"]}, {"$set": sets})
            out.append({"phash": r["phash"], "item_type": it,
                        "color": res.get("color"), "fabric": res.get("fabric_guess"),
                        "confidence": res.get("confidence")})
        return {"classified": len(out), "provider": body.provider, "results": out}

    # ── SKUs ──────────────────────────────────────────────────────────────────
    @router.get("/skus")
    async def list_skus(authorization: Optional[str] = Header(None)):
        guard(authorization)
        skus = await db.az_skus.find({}, {"_id": 0}).sort("created", -1).to_list(500)
        counts = {}
        async for row in db.az_images.aggregate([
                {"$match": {"sku": {"$ne": ""}}},
                {"$group": {"_id": "$sku",
                            "images": {"$sum": 1},
                            "slots": {"$push": "$slot"},
                            "ready": {"$sum": {"$cond": [
                                {"$and": [{"$eq": ["$qa_status", "pass"]},
                                          {"$eq": ["$uploaded", False]}]}, 1, 0]}}}}]):
            counts[row["_id"]] = row
        for s in skus:
            c = counts.get(s["sku"], {})
            s["images"] = c.get("images", 0)
            s["slots"] = sorted(x for x in c.get("slots", []) if x)
            s["ready"] = c.get("ready", 0)
        return skus

    # ── Templates ─────────────────────────────────────────────────────────────
    @router.post("/templates")
    async def upload_template(file: UploadFile = File(...),
                              name: str = Form(...),
                              product_type: str = Form(...),
                              item_types: str = Form(...),   # csv, e.g. "saree" or "dress_material,kurti"
                              authorization: Optional[str] = Header(None)):
        guard(authorization)
        if not _OPENPYXL_OK:
            raise HTTPException(status_code=500, detail="openpyxl not installed on server")
        if not (file.filename or "").lower().endswith((".xlsm", ".xlsx")):
            raise HTTPException(status_code=400, detail="Template must be .xlsm or .xlsx")
        tid = str(uuid.uuid4())[:8]
        ext = ".xlsm" if file.filename.lower().endswith(".xlsm") else ".xlsx"
        path = TEMPLATE_DIR / f"{tid}{ext}"
        data = await file.read()
        path.write_bytes(data)
        try:
            info = await asyncio.to_thread(_introspect_template, path)
        except HTTPException:
            path.unlink(missing_ok=True)
            raise
        doc = {
            "id": tid, "name": name.strip(), "file": path.name,
            "src_filename": file.filename,
            "product_type": product_type.strip().upper(),
            "item_types": [t.strip() for t in item_types.split(",") if t.strip()],
            "columns": info["columns"], "total_columns": info["total_columns"],
            "image_slots": [s for s in SLOTS if s in info["columns"]],
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.az_templates.insert_one({**doc})
        return doc

    @router.get("/templates")
    async def list_templates(authorization: Optional[str] = Header(None)):
        guard(authorization)
        return await db.az_templates.find({}, {"_id": 0}).sort("created_at", -1).to_list(50)

    @router.delete("/templates/{tid}")
    async def delete_template(tid: str, authorization: Optional[str] = Header(None)):
        guard(authorization)
        doc = await db.az_templates.find_one({"id": tid}, {"_id": 0})
        if not doc:
            raise HTTPException(status_code=404, detail="Template not found")
        (TEMPLATE_DIR / doc["file"]).unlink(missing_ok=True)
        await db.az_templates.delete_one({"id": tid})
        return {"deleted": tid}

    # ── Export ────────────────────────────────────────────────────────────────
    @router.post("/export")
    async def export(body: ExportRequest, authorization: Optional[str] = Header(None)):
        guard(authorization)
        tpl = await db.az_templates.find_one({"id": body.template_id}, {"_id": 0})
        if not tpl:
            raise HTTPException(status_code=404, detail="Template not found")
        q = {"uploaded": False, "qa_status": "pass", "sku": {"$ne": ""}, "slot": {"$ne": ""}}
        if body.batch_date:
            q["batch_date"] = body.batch_date
        rows = await db.az_images.find(q, {"_id": 0}).to_list(2000)
        if not rows:
            raise HTTPException(status_code=400, detail="No QA-passed, SKU-assigned images ready to export")

        by_sku: dict = {}
        for r in rows:
            by_sku.setdefault(r["sku"], []).append(r)

        items, skipped = [], []
        for sku, imgs in by_sku.items():
            meta = await db.az_skus.find_one({"sku": sku}, {"_id": 0}) or {}
            itype = meta.get("item_type", "?")
            if itype not in tpl["item_types"]:
                skipped.append({"sku": sku, "item_type": itype})
                continue
            items.append({"sku": sku, "item_name": meta.get("item_name", ""),
                          "images": {r["slot"]: r["lib_file"] for r in imgs},
                          "phashes": [r["phash"] for r in imgs]})
        if not items:
            raise HTTPException(
                status_code=400,
                detail=f"Ready SKUs exist but none match this template's item types "
                       f"({', '.join(tpl['item_types'])}). Held: "
                       f"{', '.join(s['sku'] + ':' + s['item_type'] for s in skipped[:10])}")

        cols = tpl["columns"]
        tpl_path = TEMPLATE_DIR / tpl["file"]
        stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d_%H%M")
        ext = Path(tpl["file"]).suffix
        out_name = f"{tpl['product_type']}_batch_{stamp}{ext}"
        out_path = EXPORT_DIR / out_name
        url_base = body.url_base.rstrip("/")

        def _write():
            wb = load_workbook(tpl_path, keep_vba=(ext == ".xlsm"))
            ws = wb["Template"]
            # Blank Amazon's ABC123/SHIRT example row — Seller Central rejects
            # files that still contain it. Our rows then reuse the freed row.
            for row in range(DATA_START, DATA_START + 3):
                if str(ws.cell(row=row, column=cols["sku"]).value or "").strip() == "ABC123":
                    for c in range(1, tpl.get("total_columns", 200) + 5):
                        ws.cell(row=row, column=c).value = None
            r = DATA_START
            written = 0
            for item in items:
                while ws.cell(row=r, column=cols["sku"]).value not in (None, ""):
                    r += 1   # skip Amazon's example row and any prefilled rows
                ws.cell(row=r, column=cols["sku"], value=item["sku"])
                if "product_type" in cols:
                    ws.cell(row=r, column=cols["product_type"], value=tpl["product_type"])
                if "action" in cols:
                    ws.cell(row=r, column=cols["action"], value="(Default) Create or Replace")
                if "item_name" in cols and item["item_name"]:
                    ws.cell(row=r, column=cols["item_name"], value=item["item_name"])
                for slot, lib_file in item["images"].items():
                    if slot in cols:
                        ws.cell(row=r, column=cols[slot], value=f"{url_base}/{lib_file}")
                r += 1
                written += 1
            wb.save(out_path)
            return written

        written = await asyncio.to_thread(_write)
        done = [p for item in items for p in item["phashes"]]
        await db.az_images.update_many({"phash": {"$in": done}}, {"$set": {"uploaded": True}})
        return {"file": out_name, "download_url": f"/api/uploads/amazon/exports/{out_name}",
                "skus_written": written, "images_marked": len(done),
                "skipped": skipped,
                "note": "Fill remaining required columns (price, quantity, browse node, fabric...) "
                        "per SKU in the file before uploading to Seller Central."}

    # ── Status dashboard ──────────────────────────────────────────────────────
    @router.get("/status")
    async def status(authorization: Optional[str] = Header(None)):
        guard(authorization)
        batches = []
        async for row in db.az_images.aggregate([
                {"$group": {"_id": "$batch_date", "total": {"$sum": 1},
                            "passed": {"$sum": {"$cond": [{"$eq": ["$qa_status", "pass"]}, 1, 0]}},
                            "review": {"$sum": {"$cond": [{"$eq": ["$qa_status", "manual_review"]}, 1, 0]}},
                            "failed": {"$sum": {"$cond": [{"$eq": ["$qa_status", "fail"]}, 1, 0]}},
                            "assigned": {"$sum": {"$cond": [{"$ne": ["$sku", ""]}, 1, 0]}},
                            "exported": {"$sum": {"$cond": ["$uploaded", 1, 0]}}}},
                {"$sort": {"_id": -1}}]):
            batches.append({"batch_date": row["_id"], **{k: v for k, v in row.items() if k != "_id"}})
        pending_ai = await db.az_images.count_documents(
            {"ai_checked": False, "qa_status": {"$ne": "fail"}, "lib_file": {"$ne": ""}})
        need_input = await db.az_images.count_documents(
            {"uploaded": False, "qa_status": {"$ne": "fail"},
             "$or": [{"sku": ""}, {"item_type": {"$in": ["", "?"]}}]})
        drive_ok = os.path.exists(os.environ.get("GDRIVE_SA_FILE", "/app/gdrive-sa.json"))
        return {"batches": batches, "pending_ai": pending_ai, "need_input": need_input,
                "providers": list(VISION_PROVIDERS),
                "anthropic_key_set": bool(os.environ.get("ANTHROPIC_API_KEY")),
                "drive_configured": drive_ok}

    return router
